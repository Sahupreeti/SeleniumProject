'''from selenium import webdriver
import time
from selenium.webdriver.common.by import By
'''
from openpyxl import Workbook

'''path = "C:/SeleniumDriver/chromedriver.exe"
url = "https://www.google.com"
search_val = "Flipkart"
driver = webdriver.Chrome(path)
driver.get(url)
driver.implicitly_wait(10)
driver.maximize_window()
driver.find_element(By.CSS_SELECTOR, "#yDmH0d > c-wiz > div > div > c-wiz > div > div > div > div.DRc6kd.bdn4dc > div.QlyBfb > button").click()
#driver.find_element(By.ID, "input").send_keys(search_val).click()
'''

wb = Workbook()
ws = wb.active
#ws["A1"]= "What is your name?"
'''list = [["Employee_name","Emp_No"],["Ram", "1"],["Raghu", "2"]]
for value in list:
    ws.append(value)
wb.save("newexcel.xlsx")
'''
for r in range(1, 8):
    for c in range(1, 4):
        ws.cell(row=r, column=c).value = r+c
print("values added into excel")