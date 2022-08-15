from openpyxl import Workbook

wb = Workbook()
ws = wb.active
#ws['A2'] = "Learning excel in Python"
'''my_list = [["Name", "City"], ["Ram", "Lilapur"], ["Sita", "Janakpur"], ["Laxman", "Ayodhya"]]
for data in my_list:
    ws.append(data)
'''
#saving data into excelfile
for r in range(1, 6):
    for c in range(1,5):
        ws.cell(row=r, column=c).value = r+c
wb.save("demoexcel1_file.xlsx")
