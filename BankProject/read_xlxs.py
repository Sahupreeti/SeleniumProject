from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='demoexcel_file.xlsx')
sh = wb.active
print(sh["A2"].value, sh["B2"].value)
row_ct = sh.max_row
col_ct = sh.max_column

for r in range(1, row_ct+1):
    for c in range(1, col_ct+1):
        print(sh.cell(row=r, column=c).value, end =' ')
    print("\n")